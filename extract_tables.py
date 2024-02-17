#
# extract_tables.py
# Extract tables from xls files
#

import json
from openpyxl import load_workbook
from tqdm import tqdm
from os import walk
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import logging
import os
import re
import unicodedata

logFormatter = logging.Formatter(
    "%(asctime)s [%(threadName)-12.12s] [%(levelname)-5.5s]\t%(message)s")
skippedTables_logFormatter = logging.Formatter(
    "%(message)s")

rootLogger = logging.getLogger('Main')
skippedLogger = logging.getLogger('SkippedTables')

fileHandler = logging.FileHandler(
    "{0}/{1}.log".format('./', 'output'), 'w')

skippedTables_fileHandler = logging.FileHandler(
    "{0}/{1}.log".format('./', 'skipped_tables'), 'w')

fileHandler.setFormatter(logFormatter)
skippedTables_fileHandler.setFormatter(skippedTables_logFormatter)

rootLogger.addHandler(fileHandler)
skippedLogger.addHandler(skippedTables_fileHandler)

# consoleHandler = logging.StreamHandler()
# consoleHandler.setFormatter(logFormatter)
# rootLogger.addHandler(consoleHandler)
rootLogger.setLevel(logging.INFO)
skippedLogger.setLevel(logging.INFO)


def _get_cell_font_attributes(cell):
    """
    Get the cell font attributes
    :param cell:  the input cell
    :return: a dictionary with the attributes
    """
    attrs = {
        'font_name': None,
        'font_size': None,
        'wrap_text': None,
        'BC': 0,  # non-white background
        'FC': 0,  # non-black font color
        'FB': 0,  # has font bold
        'I': 0,  # has font italic
    }
    if cell.has_style:
        font = cell.font
        alignment = cell.alignment
        attrs['font_name'] = font.name
        attrs['font_size'] = font.sz
        attrs['wrap_text'] = alignment.wrapText
        if font.b:
            attrs['FB'] = 1
        if font.i:
            attrs['I'] = 1
        color = font.color
        if color is not None:
            if color.rgb is None or color.type != 'rgb':
                attrs['FC'] = 0
            else:
                attrs['FC'] = 1
        else:
            attrs['FC'] = 0
    return attrs


def _get_cell_data_attributes(cell):
    """
    Get the data attributes of a cell
    :param cell: the input cell
    :return: a dictionary with the data attributes
    """
    attrs = {'NS': cell.number_format, 'DT': 5}
    # Get the data type
    dt = cell.data_type
    if dt == 's':
        attrs['DT'] = 0  # 0: string
    elif dt == 'd':
        attrs['DT'] = 2  # 2: date
    elif dt == 'n':
        # Check if percent
        if '%' in str(cell.value):
            attrs['DT'] = 3  # percentage
        elif '$' in str(cell.value):
            attrs['DT'] = 4  # currency
        else:
            attrs['DT'] = 1  # number
    return attrs


def _get_cell_border_attributes(cell):
    """
    Get the border attributes for the cell
    :param cell: the input cell
    :return: a dictionary with the border attributes
    """
    attrs = {
        'LB': 0,  # has left border
        'TB': 0,  # has top border
        'BB': 0,  # has bottom border
        'RB': 0  # has right border
    }
    if cell.border is not None:
        if cell.border.left.style is not None:
            attrs['LB'] = 1
        if cell.border.top.style is not None:
            attrs['TB'] = 1
        if cell.border.bottom.style is not None:
            attrs['BB'] = 1
        if cell.border.right.style is not None:
            attrs['RB'] = 1
    return attrs


def _get_cell_alignment_attributes(cell):
    """
    Get the alignment attributes of the cell

    For the horizontal alignment:
    # center=0,
    # center_across_selection=1,
    # distributed=2,
    # fill=3,
    # general=4,
    # justify=5,
    # left=6,
    # right=7
    Openpyxl offers the following:
    {fill, left, distributed, general, centerContinuous, justify, center, right}

    For the vertical alignment:
    # top=0,
    # center=1,
    # bottom=2,
    # justify=3,
    # distributed=4
    Openpyxl offers the following:
    {distributed, top, justify, center, bottom}

    :param cell: the input cell
    :return: a dictionary with the alignment attributes
    """
    attrs = {
        'O': None,  # orientation?
        'HA': 0,  # horizontal align
        'VA': 1  # vertical align
    }
    if cell.has_style:
        alignment = cell.alignment
        if alignment.horizontal == 'center':
            attrs['HA'] = 0
        if alignment.horizontal == 'centerContinuous':
            attrs['HA'] = 1
        if alignment.horizontal == 'distributed':
            attrs['HA'] = 2
        if alignment.horizontal == 'fill':
            attrs['HA'] = 3
        if alignment.horizontal == 'general':
            attrs['HA'] = 4
        if alignment.horizontal == 'justify':
            attrs['HA'] = 5
        if alignment.horizontal == 'left':
            attrs['HA'] = 6
        if alignment.horizontal == 'right':
            attrs['HA'] = 7
        if alignment.vertical == 'top':
            attrs['VA'] = 0
        if alignment.vertical == 'center':
            attrs['VA'] = 1
        if alignment.vertical == 'bottom':
            attrs['VA'] = 2
        if alignment.vertical == 'justify':
            attrs['VA'] = 3
        if alignment.vertical == 'distributed':
            attrs['VA'] = 4
        attrs['O'] = alignment.textRotation
    return attrs


def _get_merged_regions(worksheet, table_content, removed_idx, removed_rows):
    """
    Get the merged regions of the given worksheet
    :param worksheet: the input worksheet
    :return: a list with the merged regions
    """
    regions = []
    merged_ranges = worksheet.merged_cells.ranges
    for m in merged_ranges:
        if m is not None:
            start = str(m).split(':')[0]
            end = str(m).split(':')[1]
            start_xy = coordinate_from_string(start)
            end_xy = coordinate_from_string(end)
            regions.append({
                'FirstRow': start_xy[1] - 1,  # index starts at (0,0)
                'LastRow': end_xy[1] - 1,
                'FirstColumn': column_index_from_string(start_xy[0]) - 1,
                'LastColumn': column_index_from_string(end_xy[0]) - 1
            })
    # Sanity checks and corrections due to possibly removed rows
    if removed_idx is None and len(removed_rows) == 0:
        return regions
    final_regions = []
    num_table_rows = len(table_content)
    if removed_idx:
        # This means that a row has been removed and the rest of the table content has been moved up
        for region in regions:
            if region['FirstRow'] > removed_idx and region['LastRow'] > removed_idx:
                region['FirstRow'] = region['FirstRow'] - 1
                region['LastRow'] = region['LastRow'] - 1
    # Finally, keep those regions that are between table boundaries. Otherwise this means that they
    # were at the bottom (with foonotes) and they have been removed
    for region in regions:
        if region['FirstRow'] < num_table_rows and region['LastRow'] < num_table_rows:
            final_regions.append(region)
    return final_regions


def _get_top_tree(table):
    """
    Infer the top tree from the given table
    :param table: the given table in a dict representation
    :return: a dictionary with the updated table
    """
    top_header_rows_number = 1
    # All nodes in the tree contain table coordinates (and not tree coordinates)
    # Set the top tree root which is always (-1,-1)
    top_tree = {
        'RI': -1,
        'CI': -1,
        'Cd': []
    }
    # Check for the header rows
    # Scan the merged regions to see whether the top left cell is merged
    merged_regions = table['MergedRegions']
    for region in merged_regions:
        if region['FirstRow'] == 0 and region['LastRow'] == 1 \
                and region['FirstColumn'] == 0 and region['LastColumn'] == 0:
            top_header_rows_number = 2
            break
    if top_header_rows_number == 1:
        top_row = table['Cells'][0]
        for index, cell in enumerate(top_row):
            if index != 0:
                node = {
                    'RI': cell['coordinates'][0],
                    'CI': cell['coordinates'][1],
                    'Cd': []
                }
                top_tree['Cd'].append(node)

    if top_header_rows_number == 2:
        # Check merged regions for top row:
        merged_columns = []
        for region in merged_regions:
            if region['FirstRow'] == 0 and region['LastRow'] == 0:
                # get the merged columns:
                merged_columns.append((region['FirstColumn'], region['LastColumn']))
        top_row = table['Cells'][0]
        for index, cell in enumerate(top_row):
            if index != 0:
                node = {
                    'RI': cell['coordinates'][0],
                    'CI': cell['coordinates'][1],
                    'Cd': []
                }
                if cell['value'] != 'None':
                    top_tree['Cd'].append(node)
        # Get the second row
        second_row = table['Cells'][1]
        # Check if we have merged columns on top row
        if len(merged_columns) == 0:
            # Then  the first column is child of the above (row,column) and the
            # remaining columns are childer of topRoot
            for idx, cell in enumerate(second_row):
                if idx == 1:
                    node = {
                        'RI': cell['coordinates'][0],
                        'CI': cell['coordinates'][1],
                        'Cd': []
                    }
                    if cell['value'] != 'None':
                        top_tree['Cd'][0]['Cd'].append(node)
                if idx > 1:
                    node = {
                        'RI': cell['coordinates'][0],
                        'CI': cell['coordinates'][1],
                        'Cd': []
                    }
                    if cell['value'] != 'None':
                        top_tree['Cd'].append(node)
        else:
            # There are merged regions in the top row
            top_nodes_number = len(merged_columns)
            for index, cell in enumerate(second_row):
                node_added = False
                if index != 0:
                    node = {
                        'RI': cell['coordinates'][0],
                        'CI': cell['coordinates'][1],
                        'Cd': []
                    }
                    for idx_merged, top_merges in enumerate(merged_columns):
                        if top_merges[0] <= node['CI'] <= top_merges[1]:
                            parent_node = idx_merged
                            if parent_node < len(top_tree['Cd']):
                                # This means that there is a parent node for this node
                                top_tree['Cd'][parent_node]['Cd'].append(node)
                            else:
                                # Else although there is a merged region, it is by mistake and
                                # has no value
                                # (see: 888491_2020_10-K_0000888491-20-000007.xlsx -> SUMMARY OF SIGNIFICANT ACCOUNTING POLICIES (Narrative) (Detail))
                                top_tree['Cd'].append(node)
                            node_added = True
                            break
                    if node_added is False:
                        # Then this node does not belong under a merged region
                        # and should be a direct child of the root node
                        top_tree['Cd'].append(node)
    # We always need to include the first cell of the row as a child of top tree in the first place
    # Get the col coordinate of the first child that we have included up to now:
    row_idx = top_tree['Cd'][0]['RI']
    col_idx = top_tree['Cd'][0]['CI']
    for i in reversed(range(0, col_idx)):
        node = {
            'RI': row_idx,
            'CI': i,
            'Cd': []
        }
        top_tree['Cd'].insert(0, node)
    updated_table = {
        'TopHeaderRowsNumber': top_header_rows_number,
        'TopTreeRoot': top_tree
    }
    return updated_table


def _get_left_tree(table):
    """
    Infer the left tree from the given table
    :param table: the given table in a dict representation
    :return: an updated table with the left tree info
    """
    left_header_columns_number = 1
    # All nodes in the tree contain table coordinates (and not tree coordinates)
    # Set the left tree root which is always (-1,-1)
    left_tree = {
        'RI': -1,
        'CI': -1,
        'Cd': []
    }
    # Get the top header rows number to see in which row we should start
    top_headers = table['TopHeaderRowsNumber']
    row_start = top_headers
    # Scan the rows from that starting point
    have_seen_bolds = False
    parent = -1
    for row_number in range(row_start, len(table['Cells'])):
        row = table['Cells'][row_number]
        # Decide if this is a top child
        # Check if the cell is bold
        cell = row[0]
        node = {
            'RI': cell['coordinates'][0],
            'CI': cell['coordinates'][1],
            'Cd': []
        }
        if cell['FB'] == 1:
            left_tree['Cd'].append(node)
            have_seen_bolds = True
            left_header_columns_number = 2
            parent += 1
        else:
            if have_seen_bolds is False:
                left_tree['Cd'].append(node)
                parent += 1
            else:
                left_tree['Cd'][parent]['Cd'].append(node)
    # We always need to include the first cell of the column as a child of left tree in the first place
    # Get the row coordinate of the first child that we have included up to now:
    row_idx = left_tree['Cd'][0]['RI']
    col_idx = left_tree['Cd'][0]['CI']
    for i in reversed(range(0, row_idx)):
        node = {
            'RI': i,
            'CI': col_idx,
            'Cd': []
        }
        left_tree['Cd'].insert(0, node)
    updated_table = {
        'LeftHeaderColumnsNumber': left_header_columns_number,
        'LeftTreeRoot': left_tree
    }
    return updated_table


def _calculate_dimensions(original_dims, table_content):
    """
    Calculate the dimensions of the table given the original dimensions
    and the final content (i.e. table content)
    :param original_dims: the original dimensions
    :param table_content: the table content
    :return: the new dimensions
    """
    num_rows_in_content = len(table_content)
    final_dims = original_dims
    original_start = original_dims.split(':')[0]
    original_end = original_dims.split(':')[1]
    match_start = re.match(r"([a-z]+)([0-9]+)", original_start, re.I)
    match_end = re.match(r"([a-z]+)([0-9]+)", original_end, re.I)
    if match_start and match_end:
        start_items = match_start.groups()
        start_row = int(start_items[1])
        start_col = start_items[0]
        end_items = match_end.groups()
        end_row = int(end_items[1])
        end_col = end_items[0]
        # Check if there is a missmatch because of lines that have been removed
        if (end_row - start_row + 1) == num_rows_in_content:
            return final_dims
        else:
            diff = end_row - num_rows_in_content
            end_row = end_row - diff
            # create final dims
            final_dims = start_col + str(start_row) + ':' + end_col + str(end_row)
            return final_dims


def process_ws(ws):
    """
    Process the specified worksheet
    :param ws: The worksheet to be processed
    :return: The table of the worksheet
    """
    # Access the table data based on the sheet dimensions
    data = ws[ws.dimensions]
    content = [[cell.value for cell in ent]
               for ent in data
               ]
    styles = []
    for ent in data:
        cell_styles = []
        for cell in ent:
            style = {
                'HF': 0,  # has_formula
                'A1': '',  # formula-specific
                'R1': '',  # formula-specific
            }
            font_attrs = _get_cell_font_attributes(cell)
            data_atts = _get_cell_data_attributes(cell)
            border_attrs = _get_cell_border_attributes(cell)
            alignment_attrs = _get_cell_alignment_attributes(cell)
            # Other styles to consider in the future: cell.fill, cell.protection
            style.update(font_attrs)
            style.update(data_atts)
            style.update(border_attrs)
            style.update(alignment_attrs)
            cell_styles.append(style)
        styles.append(cell_styles)

    # Do not process the worksheet if there is a tiny table
    # if len(content) < 5:
    #    skippedLogger.info(f'{ws.title}')
    #    return None
    # Do not process the worksheet if there are many empty rows
    num_empty_rows = 0
    num_columns = len(content[0])
    empty_row_idx = None
    for rowid, row in enumerate(content):
        num_nones = 0
        for column in row:
            if column is None or unicodedata.normalize('NFKD', str(column)).strip() == "" or unicodedata.normalize(
                    'NFKD', str(column)).strip() == "None":
                num_nones += 1
        if num_nones == num_columns:
            num_empty_rows += 1
            empty_row_idx = rowid
    if num_empty_rows > 1:
        # Don't process the table
        return None
    if 0 < num_empty_rows < 2 and empty_row_idx is not None:
        content.pop(empty_row_idx)
        styles.pop(empty_row_idx)

    # check if the last line of the table is footnote:
    rows_with_footnotes = []
    for i in reversed(range(len(content))):
        matches = []
        if content[i][0] is not None:
            matches = re.findall(r"\[\d\]", content[i][0])
        if len(matches) > 0:
            rows_with_footnotes.append(i)
    # remove rows with footnotes
    for index in rows_with_footnotes:
        content.pop(index)
        styles.pop(index)

    # Do not process the worksheet if there are cells with large text content
    # Scan the content to see if such cells exist
    for row in content:
        for cell in row:
            cell_value = str(cell)
            if len(cell_value.split()) > 20:
                skippedLogger.info(f'{ws.title}')
                return None
    # Get table title this is the cell (0,0), otherwise the spreadsheet name
    title = str(content[0][0])
    if title is None or title == '' or title == ' ':
        title = ws.title
    table_range = _calculate_dimensions(ws.dimensions, content)
    table = {
        'StorageAccount': 'EDGARExcelCrawled',
        'BlobName': 'DataSpreadsheet',
        'SheetName': ws.title,
        'Language': 'english',
        'RangeAddress': table_range,
        'Title': title,
        'Cells': []
    }
    # Process content
    for idx, row in enumerate(content):
        is_header = False
        # Assumption about the first row
        if idx == 0:
            is_header = True

        # Heuristics to assess whether a row seems like a header
        if row[0] is None:
            is_header = True
        num_of_nones = sum(x is None for x in row)
        if num_of_nones == len(row) - 1:
            is_header = True

        # Create cell info for this row
        cells = []
        styles_for_this_row = styles[idx]
        for i, d in enumerate(row):
            is_attribute = False
            if not is_header:
                if d is not None:
                    if i == 0:
                        is_attribute = True
            if str(d) == "None":
                v = ""
            else:
                v = str(d)
            cell = {
                'T': str(d),  # cell text
                'V': unicodedata.normalize('NFKD', v).strip(),
                'is_header': is_header,
                'value': str(d),
                'is_attribute': is_attribute,
                'coordinates': (idx, i)  # (x,y)
            }
            cell.update(styles_for_this_row[i])
            cells.append(cell)
        table['Cells'].append(cells)
    # Get the merged regions
    merged_regions = _get_merged_regions(ws, content, empty_row_idx, rows_with_footnotes)
    table['MergedRegions'] = merged_regions
    # Get the trees
    top_tree_info = _get_top_tree(table)
    table.update(top_tree_info)
    left_tree_info = _get_left_tree(table)
    table.update(left_tree_info)
    return table


def process_wb(wb):
    """
    Process the sheets of the specified workbook
    :param wb: The workbook to be processed
    :return: A list with extracted tables as dictionaries
    """
    # Get a list with all the worksheets
    worksheets = wb.worksheets
    # Process the worksheets that have meaningful information
    extracted_tables = []
    for i in range(len(worksheets)):
        try:
            worksheet = wb[worksheets[i].title]
            json_table = process_ws(worksheet)
            if json_table is not None:
                extracted_tables.append(json_table)
        except:
            rootLogger.error(f'Skipped sheet: {worksheets[i].title}')
    return extracted_tables


def batch_process_wb(directory):
    """
    Batch processing of workbooks in the specified directory
    :param directory: The dirrectory containing the workbooks
    :return:
    """
    print("Getting filenames..")
    files = []
    for (dirpath, dirnames, filenames) in walk(directory):
        for file in filenames:
            files.append(os.path.join(dirpath, file))
    cnt_errors = 0
    print('Processing workbooks..')
    for file in tqdm(files):
        try:
            workbook = load_workbook(file)
            wb_tables = process_wb(workbook)
            rootLogger.info(f'Processed file: {file}: Found {len(wb_tables)} tables.')
            # Save the extracted tables to output directory
            output_filename = './output/' + file.split('/')[-1].split('.')[0] + '.json'
            with open(output_filename, 'w') as fp:
                fp.write(json.dumps(wb_tables))
        except InvalidFileException:
            cnt_errors += 1
            rootLogger.error(f'Skipped file: {file}')
        except:
            cnt_errors += 1
            rootLogger.error(f'Skipped file: {file}')
    print(f'Skipped {cnt_errors} files.')


if __name__ == "__main__":
    # batch_process_wb('/Users/Ilias.Zavitsanos@ey.com/DATASETS/excel_files/10-K')
    batch_process_wb('./temp_data')
