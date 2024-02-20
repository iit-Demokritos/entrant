#
# all_tables_from_10Q_test.py
# Test the table extraction process from a 10Q report
#

import sys
import unittest
from extract_tables import process_wb
from openpyxl import load_workbook
import os

sys.path.append('../')


class TestWbTableExtraction10Q(unittest.TestCase):
    """Test the table extraction process"""

    def setUp(self):
        # Load a couple of test 10-Q report and extract data
        ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
        workbook = load_workbook(ROOT_DIR + '/test-data/10-Q.xlsx')
        self.tables_10Q = process_wb(workbook)
        workbook = load_workbook(ROOT_DIR + '/test-data/10-Q_sample2.xlsx')
        self.tables_10Q_2 = process_wb(workbook)

    def test_number_of_tables(self):
        """
        Test the number of extracted tables
        :return:
        """
        # Should extract at least 7 tables
        self.assertGreaterEqual(len(self.tables_10Q), 7)
        self.assertGreaterEqual(len(self.tables_10Q_2), 7)

    def test_table_titles(self):
        """
        Test the title for each table
        :return:
        """
        # Should have the following titles for the first 7 tables
        titles = [
            'Cover Page - shares shares in Thousands',
            'CONDENSED CONSOLIDATED STATEMENTS OF OPERATIONS (Unaudited) - USD ($) shares in Thousands, $ in Millions',
            'CONDENSED CONSOLIDATED STATEMENTS OF COMPREHENSIVE INCOME (Unaudited) - USD ($) $ in Millions',
            'CONDENSED CONSOLIDATED BALANCE SHEETS (Unaudited) - USD ($) $ in Millions',
            'CONDENSED CONSOLIDATED BALANCE SHEETS (Unaudited) (Parenthetical) - $ / shares',
            'CONDENSED CONSOLIDATED STATEMENTS OF SHAREHOLDERS\' EQUITY (Unaudited) - USD ($) $ in Millions',
            'CONDENSED CONSOLIDATED STATEMENTS OF CASH FLOWS (Unaudited) - USD ($) $ in Millions',

        ]
        for i, t in enumerate(titles):
            self.assertEqual(self.tables_10Q[i]['Title'], t)
        # For the 2nd document:
        titles = [
            'Document and Entity Information - $ / shares',
            'INCOME STATEMENTS - USD ($) shares in Millions, $ in Millions',
            'COMPREHENSIVE INCOME STATEMENTS - USD ($) $ in Millions',
            'BALANCE SHEETS - USD ($) $ in Millions',
            'BALANCE SHEETS (Parenthetical) - USD ($) $ in Millions',
            'CASH FLOWS STATEMENTS - USD ($) $ in Millions',
            'STOCKHOLDERS\' EQUITY STATEMENTS - USD ($) $ in Millions'
        ]
        for i, t in enumerate(titles):
            self.assertEqual(self.tables_10Q_2[i]['Title'], t)

    def test_table_rows_lengths(self):
        """
        Test the rows for each table
        :return:
        """
        # Should have the following rows for each table
        lengths = [84, 26, 16, 34, 6, 49, 44]
        for i, s in enumerate(lengths):
            self.assertEqual(len(self.tables_10Q[i]['Cells']), s)
        # For the 2nd document:
        lengths = [45, 25, 10, 38, 6, 39, 23]
        for i, s in enumerate(lengths):
            self.assertEqual(len(self.tables_10Q_2[i]['Cells']), s)

    def test_table_headers(self):
        """
        Test the headers for each table
        :return:
        """
        # Should the following rows for headers
        headers = [
            [0, 1, 2, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80],
            [0, 1, 5, 14, 17, 20, 23],
            [0, 1, 2, 4, 6, 10],
            [0, 1, 9, 15, 22, 28],
            [0, 1],
            [0, 2, 11, 14, 23, 26, 35, 38, 47],
            [0, 1, 2, 4, 6, 11, 20, 28, 41]
        ]
        for i, table_headers in enumerate(headers):
            for h in table_headers:
                self.assertEqual(self.tables_10Q[i]['Cells'][h][0]['is_header'], True)
        # For the 2nd document:
        headers = [
            [0, 1, 2, 31, 36, 41],
            [0, 1, 13, 16],
            [0, 1, 2, 4],
            [0, 1, 16, 32],
            [0, 1],
            [0, 1, 2, 4, 9, 20, 27],
            [0]
        ]
        for i, table_headers in enumerate(headers):
            for h in table_headers:
                self.assertEqual(self.tables_10Q_2[i]['Cells'][h][0]['is_header'], True)

    def test_table_cells_lengths(self):
        """
        Test the cells for each table
        :return:
        """
        # Should have the following cells
        cells = [3, 5, 5, 3, 3, 7, 3]
        for i, cell in enumerate(cells):
            for row in self.tables_10Q[i]['Cells']:
                self.assertEqual(len(row), cell)
        # For the 2nd document
        cells = [3, 3, 3, 3, 3, 3, 7]
        for i, cell in enumerate(cells):
            for row in self.tables_10Q_2[i]['Cells']:
                self.assertEqual(len(row), cell)


suite = unittest.TestLoader().loadTestsFromTestCase(TestWbTableExtraction10Q)
unittest.TextTestRunner(verbosity=2).run(suite)
