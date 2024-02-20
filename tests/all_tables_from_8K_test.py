#
# all_tables_from_8K_test.py
# Test the table extraction process from a 8K report
#

import sys
import unittest
from extract_tables import process_wb
from openpyxl import load_workbook
import os

sys.path.append('../')


class TestWbTableExtraction8K(unittest.TestCase):
    """Test the table extraction process"""

    def setUp(self):
        # Load a couple of test 8-K report and extract data
        ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
        workbook = load_workbook(ROOT_DIR + '/test-data/8-K.xlsx')
        self.tables_8K = process_wb(workbook)
        workbook = load_workbook(ROOT_DIR + '/test-data/8-K_sample2.xlsx')
        self.tables_8K_2 = process_wb(workbook)

    def test_number_of_tables(self):
        """
        Test the number of extracted tables
        :return:
        """
        # Should extract at least 1 table
        self.assertGreaterEqual(len(self.tables_8K), 1)
        self.assertGreaterEqual(len(self.tables_8K_2), 1)

    def test_table_titles(self):
        """
        Test the title for each table
        :return:
        """
        # Should have the following titles for the first table
        titles = [
            'Cover'
        ]
        for i, t in enumerate(titles):
            self.assertEqual(self.tables_8K[i]['Title'], t)
        # For the 2nd document:
        titles = [
            'Document and Entity Information'
        ]
        for i, t in enumerate(titles):
            self.assertEqual(self.tables_8K_2[i]['Title'], t)

    def test_table_rows_lengths(self):
        """
        Test the rows for each table
        :return:
        """
        # Should have the following rows for each table
        lengths = [131]
        for i, s in enumerate(lengths):
            self.assertEqual(len(self.tables_8K[i]['Cells']), s)
        # For the 2nd document:
        lengths = [36]
        for i, s in enumerate(lengths):
            self.assertEqual(len(self.tables_8K_2[i]['Cells']), s)

    def test_table_headers(self):
        """
        Test the headers for each table
        :return:
        """
        # Should the following rows for headers
        headers = [
            [0, 1, 22, 27, 32, 37, 42, 47, 52, 57, 62, 67, 72, 77, 82, 87, 92, 97, 102, 107, 112, 117, 122,
             127]
        ]
        for i, table_headers in enumerate(headers):
            for h in table_headers:
                self.assertEqual(self.tables_8K[i]['Cells'][h][0]['is_header'], True)
        # For the 2nd document:
        headers = [
            [0, 1, 22, 27, 32]
        ]
        for i, table_headers in enumerate(headers):
            for h in table_headers:
                self.assertEqual(self.tables_8K_2[i]['Cells'][h][0]['is_header'], True)

    def test_table_cells_lengths(self):
        """
        Test the cells for each table
        :return:
        """
        # Should have the following cells
        cells = [2]
        for i, cell in enumerate(cells):
            for row in self.tables_8K[i]['Cells']:
                self.assertEqual(len(row), cell)
        # For the 2nd document:
        for i, cell in enumerate(cells):
            for row in self.tables_8K_2[i]['Cells']:
                self.assertEqual(len(row), cell)


suite = unittest.TestLoader().loadTestsFromTestCase(TestWbTableExtraction8K)
unittest.TextTestRunner(verbosity=2).run(suite)
