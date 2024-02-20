#
# table_from_10Q_test.py
# Test the table extraction process from a 10Q report
#

import sys
import unittest
from extract_tables import process_ws
from openpyxl import load_workbook
import os

sys.path.append('../')


class TestTableExtraction10Q(unittest.TestCase):
    """Test the table extraction process"""

    def setUp(self):
        # Load a couple of test 10-Q report and extract data
        ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
        workbook = load_workbook(ROOT_DIR + '/test-data/10-Q.xlsx')
        worksheets = workbook.worksheets
        worksheet = workbook[worksheets[0].title]
        self.table_10Q = process_ws(worksheet)
        workbook = load_workbook(ROOT_DIR + '/test-data/10-Q_sample2.xlsx')
        worksheets = workbook.worksheets
        worksheet = workbook[worksheets[0].title]
        self.table_10Q_2 = process_ws(worksheet)

    def test_table_title(self):
        """
        Test the table title
        :return:
        """
        # Should have title "Cover Page" and "Document and Entity Information for the 2nd doc
        self.assertEqual(self.table_10Q['Title'], 'Cover Page - shares shares in Thousands')
        self.assertEqual(self.table_10Q_2['Title'], 'Document and Entity Information - $ / shares')

    def test_table_rows_len(self):
        """
        Test the table rows
        :return:
        """
        # Should have 84 rows the 10-Q table and 45 for the 2nd doc
        self.assertEqual(len(self.table_10Q['Cells']), 84)
        self.assertEqual(len(self.table_10Q_2['Cells']), 45)

    def test_table_header(self):
        """
        Test the table header
        :return:
        """
        # Should have rows 0,1,2,30,35,40,45,50,55,60,65,70,75,80 as headers
        headers = [0, 1, 2, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80]
        for i in headers:
            self.assertEqual(self.table_10Q['Cells'][i][0]['is_header'], True)
        # For the 2nd document:
        headers = [0, 1, 2, 31, 36, 41]
        for i in headers:
            self.assertEqual(self.table_10Q_2['Cells'][i][0]['is_header'], True)

    def test_table_cells_len(self):
        """
        Test the cells
        :return:
        """
        # Should have 3 cells at each row
        for row in self.table_10Q['Cells']:
            cells = row
            self.assertEqual(len(cells), 3)
        # Same for the 2nd document:
        for row in self.table_10Q_2['Cells']:
            cells = row
            self.assertEqual(len(cells), 3)


suite = unittest.TestLoader().loadTestsFromTestCase(TestTableExtraction10Q)
unittest.TextTestRunner(verbosity=2).run(suite)
