#
# table_from_8K_test.py
# Test the table extraction process from a 8K report
#

import sys
import unittest
from extract_tables import process_ws
from openpyxl import load_workbook
import os

sys.path.append('../')


class TestTableExtraction8K(unittest.TestCase):
    """Test the table extraction process"""

    def setUp(self):
        # Load a couple of test 8-K report and extract data
        ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
        workbook = load_workbook(ROOT_DIR + '/test-data/8-K.xlsx')
        worksheets = workbook.worksheets
        worksheet = workbook[worksheets[0].title]
        self.table_8K = process_ws(worksheet)
        workbook = load_workbook(ROOT_DIR + '/test-data/8-K_sample2.xlsx')
        worksheets = workbook.worksheets
        worksheet = workbook[worksheets[0].title]
        self.table_8K_2 = process_ws(worksheet)

    def test_table_title(self):
        """
        Test the table title
        :return:
        """
        # Should have title "Cover" and "Document and Entity information" for the 2nd doc
        self.assertEqual(self.table_8K['Title'], 'Cover')
        self.assertEqual(self.table_8K_2['Title'], 'Document and Entity Information')

    def test_table_rows_len(self):
        """
        Test the table rows
        :return:
        """
        # Should have 131 rows the 8-K table and 36 for the 2nd doc
        self.assertEqual(len(self.table_8K['Cells']), 131)
        self.assertEqual(len(self.table_8K_2['Cells']), 36)

    def test_table_header(self):
        """
        Test the table header
        :return:
        """
        # Rows 0,1,22,27,32,37,42,47,52,57,62,67,72,77,82,87,92,97,102,107,112,117,122,127 as headers
        headers = [0, 1, 22, 27, 32, 37, 42, 47, 52, 57, 62, 67, 72, 77, 82, 87, 92, 97, 102, 107, 112, 117,
                   122, 127]
        for i in headers:
            self.assertEqual(self.table_8K['Cells'][i][0]['is_header'], True)
        # For the 2nd document:
        headers = [0, 1, 22, 27, 32]
        for i in headers:
            self.assertEqual(self.table_8K_2['Cells'][i][0]['is_header'], True)

    def test_table_cells_len(self):
        """
        Test the cells
        :return:
        """
        # Should have 2 cells at each row
        for row in self.table_8K['Cells']:
            cells = row
            self.assertEqual(len(cells), 2)
        # For the 2nd document:
        for row in self.table_8K_2['Cells']:
            cells = row
            self.assertEqual(len(cells), 2)


suite = unittest.TestLoader().loadTestsFromTestCase(TestTableExtraction8K)
unittest.TextTestRunner(verbosity=2).run(suite)
